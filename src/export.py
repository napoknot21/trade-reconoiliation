from __future__ import annotations

import os
import json
import datetime as dt
from typing import Dict, List, Optional, Union, Any

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from src.utils import date_to_str

DFList = Union[pl.DataFrame, List[pl.DataFrame], None]
TradesByDate = Dict[dt.date, Dict[str, Dict[str, DFList]]]

def as_list(x: Any) -> List[pl.DataFrame] :
    """
    Docstring for as_list
    
    :param x: Description
    :type x: Any
    :return: Description
    :rtype: List[DataFrame]
    """
    if x is None:
        return []
    
    if isinstance(x, pl.DataFrame):
        return [x]
    
    if isinstance(x, list):
        return [d for d in x if isinstance(d, pl.DataFrame)]
    
    return []


def write_df (ws, df: pl.DataFrame, start_row: int, start_col: int = 1) -> int :
    """
    Write Polars DF at start_row; return next free row.
    """
    if df is None or df.is_empty() :
        
        ws.cell(row=start_row, column=start_col, value="(Empty table - No data yet)")
        return start_row + 1

    # header
    for j, col in enumerate(df.columns, start=start_col) :

        c = ws.cell(row=start_row, column=j, value=col)
        c.font = Font(bold=True)

    # body
    r = start_row + 1
    
    for row_vals in df.rows() :

        for j, val in enumerate(row_vals, start=start_col) :
            ws.cell(row=r, column=j, value=val)
        
        r += 1

    return r


def export_trade_reconciliation (
        
        *,
        trades_by_date: TradesByDate,
        asked_dates: List[dt.date],
        output_dir: str,
        format : str = "%Y-%m-%dT%H-%M",
    
    ) -> str:
    """
    Export reconciliation workbook from:
      trades_by_date[date][fundation][counterparty] -> list[pl.DataFrame] (or DF/None)

    Workbook:
      - one sheet per fundation
      - per date block: title + counterparty sections
    """
    os.makedirs(output_dir, exist_ok=True)
    now = date_to_str(dt.datetime.now(), format=format)

    out_path = os.path.join(output_dir, f"trade-reconciliation_{now}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    # Discover all fundations present (across dates)
    fundations: set[str] = set()

    for d in asked_dates :
        fundations |= set((trades_by_date.get(d) or {}).keys())

    for fundation in sorted(fundations) :

        ws = wb.create_sheet(title=str(fundation)[:31])
        row = 1

        for d in asked_dates :
            
            block = (trades_by_date.get(d) or {}).get(fundation) or {}

            # Title for date
            title = f"Counterparty Reconciliation at {d.isoformat()}"
            
            tcell = ws.cell(row=row, column=1, value=title)
            tcell.font = Font(bold=True, size=14)
            tcell.alignment = Alignment(horizontal="left")
            
            row += 2

            if not block :
            
                ws.cell(row=row, column=1, value="(no counterparties)")
                row += 3
            
                continue

            for cpty in sorted(block.keys()) :

                # Counterparty label
                ccell = ws.cell(row=row, column=1, value=str(cpty))
                ccell.font = Font(bold=True, size=12)
                
                row += 2  # blank row

                dfs = as_list(block.get(cpty))
                
                if not dfs :

                    ws.cell(row=row, column=1, value="(no data)")
                    row += 1
                
                else :

                    for df in dfs :

                        row = write_df(ws, df, start_row=row, start_col=1)
                        row += 1  # 1 blank row between dfs

                row += 2  # 2 blank rows between counterparties

            row += 1  # spacer between date blocks

    wb.save(out_path)

    return out_path



def save_trades_by_date_parquet (
    
        trades_by_date : dict,
        out_dir_abs_path : Optional[str] = None,
    
    ) -> Optional[str] :
    """
    Docstring for save_trades_by_date_parquet
    
    :param trades_by_date: Description
    :type trades_by_date: dict
    :param out_dir_abs_path: Description
    :type out_dir_abs_path: Optional[str]
    :return: Description
    :rtype: str | None
    """
    os.makedirs(out_dir_abs_path, exist_ok=True)

    index = {
        
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
        "format": "parquet+index_v1",
        "items": []
    
    }

    for d, fund_map in trades_by_date.items() :

        d_str = date_to_str(d)

        for fund, cpty_map in fund_map.items() :

            for cpty, dfs in cpty_map.items() :

                # dfs peut être DF ou list[DF] ou None
                if isinstance(dfs, pl.DataFrame) :
                    dfs_list = [dfs]
                
                else :
                    dfs_list = dfs or []

                for i, df in enumerate(dfs_list) :

                    if not isinstance(df, pl.DataFrame):
                        continue

                    rel_path = os.path.join(d_str, fund, cpty, f"df_{i}.parquet")
                    abs_path = os.path.join(out_dir_abs_path, rel_path)
                    
                    os.makedirs(os.path.dirname(abs_path), exist_ok=True)

                    df.write_parquet(abs_path)

                    index["items"].append(
                    
                        {
                            "date": d_str,
                            "fundation": fund,
                            "counterparty": cpty,
                            "i": i,
                            "path": rel_path,
                        }
                    
                    )

    index_path = os.path.join(out_dir_abs_path, "trades_index.json")
    
    with open(index_path, "w", encoding="utf-8") as f :
        json.dump(index, f, indent=2, ensure_ascii=False)

    return index_path