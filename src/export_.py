from __future__ import annotations

import os
import polars as pl
import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from typing import Dict, List, Optional, Union, Any

from src.utils import date_to_str

TradesType = Dict[str | dt.date, Dict[str, Union[pl.DataFrame, List[pl.DataFrame]]]]

def as_list(x: Any) -> List[pl.DataFrame] :
    """
    Docstring for _as_list
    
    :param x: Description
    :type x: Any
    :return: Description
    :rtype: List[DataFrame]
    """
    if x is None :
        return []
    
    if isinstance(x, pl.DataFrame) :
        return [x]
    
    if isinstance(x, list) :
        return [d for d in x if isinstance(d, pl.DataFrame)]
    
    return []


def write_df (ws, df : pl.DataFrame, start_row : int, start_col : int = 1) -> int:
    """
    Write df at start_row, returns next free row.
    """
    if df is None or df.is_empty() :

        ws.cell(row=start_row, column=start_col, value="(empty)")
        return start_row + 1

    # header
    for j, col in enumerate(df.columns, start=start_col) :

        c = ws.cell(row=start_row, column=j, value=col)
        c.font = Font(bold=True)

    # body
    r = start_row + 1

    for row_vals in df.rows() :

        for j, val in enumerate(row_vals, start=start_col) :
            ws.cell(row=r, column=j, value=val)
        
        r += 1

    return r


def export_trade_reconciliation (
        
        *,
        trades: TradesType,
        asked_dates: List[dt.date],
        output_dir: str,
        execution_date: Optional[dt.date] = None,
        format : str = "%Y-%m-%dT%H-%M",

    ) -> str:
    """
    Create Excel:
      - file name: trade-reconciliation_{today}.xlsx
      - sheet per fundation
      - for each asked_date:
          title
          for each counterparty:
              name
              blank row
              df1
              blank row
              df2
              ...
              2 blank rows
    """
    os.makedirs(output_dir, exist_ok=True)
    now = date_to_str(dt.datetime.now(), format=format)

    out_path = os.path.join(output_dir, f"trade-reconciliation_{now}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    for fundation, cpty_map in trades.items() :

        ws = wb.create_sheet(title=str(fundation)[:31])
        row = 1

        for d in asked_dates :
        
            # Title
            title = f"Counterparty Reconciliation at {d.isoformat()}"
            
            tcell = ws.cell(row=row, column=1, value=title)
            tcell.font = Font(bold=True, size=14)
            tcell.alignment = Alignment(horizontal="left")
            
            row += 2  # 1 blank row after title

            for cpty in sorted(cpty_map.keys()) :

                # Counterparty name
                ccell = ws.cell(row=row, column=1, value=str(cpty))
                ccell.font = Font(bold=True, size=12)
                
                row += 2  # 1 blank row

                dfs = as_list(cpty_map.get(cpty))

                if not dfs :
                
                    ws.cell(row=row, column=1, value="(no data)")
                    row += 1
                
                else :

                    for df in dfs :
                    
                        row = write_df(ws, df, start_row=row, start_col=1)
                        row += 1  # 1 blank row between dfs

                row += 2  # 2 blank rows between counterparties

            row += 1  # small spacer between dates blocks

    wb.save(out_path)

    return out_path
